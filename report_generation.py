import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
import smtplib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import logging
from config import (  # Importing email configurations
    EMAIL_FROM, EMAIL_TO, SMTP_SERVER, SMTP_USERNAME, SMTP_PASSWORD, SMTP_PORT, LOG_DIR
)
from datetime import timedelta
import openai
import re
import json
from email.mime.application import MIMEApplication
import traceback
import glob
from api_clients import api_call_with_retries
from tabulate import tabulate
import time
import pandas as pd
import json
import logging
import openai
import re
from multiprocessing import Pool, cpu_count


def generate_html_report_with_recommendations(report_entries, digest_summary, gpt_recommendations):
    """
    Generates an HTML report with summaries from the report entries, GPT-4o recommendations, and a plot of the top coins.

    Args:
        report_entries (list): List of report entries to include in the report.
        digest_summary (dict): Summary of the Sundown Digest to include at the top.
        gpt_recommendations (dict): GPT-4o's recommendations for coin purchases, structured as a list of dictionaries.
        plot_image_path (str): Path to the plot image to embed in the report.

    Returns:
        str: HTML content of the report.
    """
    
    # Sundown Digest Summary section
    digest_items = ''.join(f'<li style="font-size:14px;line-height:1.6;">{item}</li>' 
                           for item in digest_summary.get('surge_summary', [])) if digest_summary else ''
    tickers = ', '.join(digest_summary.get('tickers', [])) if digest_summary else 'N/A'

    digest_html = f"""
    <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
        <tr>
            <td style="padding:20px;">
                <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">Sundown Digest Summary</h3>
                <p style="font-size:14px;line-height:1.6;"><strong>Tickers Mentioned:</strong> {tickers}</p>
                <p style="font-size:14px;line-height:1.6;"><strong>News Summary:</strong></p>
                <ul style="list-style-type:disc;padding-left:20px;margin:0;">
                    {digest_items}
                </ul>
            </td>
        </tr>
    </table>
    """

    # Color Explanation
    color_explanation = """
    <p style="font-size:14px;line-height:1.6;">
        <strong>Color Meaning:</strong><br>
        <span style="background-color:#d4edda;padding:2px 5px;border-radius:3px;">Green</span>: Indicates coins expected to surge or break out.<br>
        <span style="background-color:#ffe5b4;padding:2px 5px;border-radius:3px;">Orange</span>: Indicates coins not expected to surge.
    </p>
    """

    # AI Recommendations Section
    if not gpt_recommendations or not gpt_recommendations.get('recommendations'):
        recommendations_html = """
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">AI Generated Coin Recommendations</h3>
                    <p style="font-size:14px;line-height:1.6;">No coins are currently recommended for purchase based on the analysis.</p>
                </td>
            </tr>
        </table>
        """
        plot_html = ""  # No plot if no recommendations
    else:
        
        # Sort recommendations by descending cumulative score (as float)
        sorted_recommendations = sorted(
            gpt_recommendations['recommendations'],
            key=lambda x: float(x.get('cumulative_score', 0)),
            reverse=True
        )

        recommendation_items = ''
        for item in sorted_recommendations:

            # Match the coin with report entries to fetch URL, cumulative score percentage
            matching_entry = next((entry for entry in report_entries if entry["coin_name"].lower() == item["coin"].lower()), None)
            
            # CoinPaprika URL format or other URL source can be used here
            coin_url = f"https://coinpaprika.com/coin/{matching_entry['coin_id']}/" if matching_entry else '#'
            cumulative_score_percentage = matching_entry.get('cumulative_score_percentage', 'N/A') if matching_entry else 'N/A'

            # Determine background color based on expected surge status
            background_color = "#d4edda" if item.get("recommendation", '').strip().lower() == "yes" else "#ffe5b4"
            
            # Capitalize each word in the coin name
            coin_name = item["coin"].title()

            recommendation_items += f"""
            <li style="font-size:14px;line-height:1.6;margin-bottom:10px;background-color:{background_color};padding:10px;border-radius:5px;">
                <b>{coin_name}</b> - {item["reason"]}<br>
                <strong>Cumulative Score Percentage:</strong> {cumulative_score_percentage}%<br>
                <a href="{coin_url}" target="_blank" style="color:#0077cc;text-decoration:none;">More Info</a>
            </li>
            """
        recommendations_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">AI Generated Coin Recommendations</h3>
                    {color_explanation}
                    <p style="font-size:14px;line-height:1.6;"><strong>Meaning of Cumulative Score Percentage:</strong> a higher percentage indicates a stronger potential based on historical data and analysis.</p>
                    <ul style="list-style-type:disc;padding-left:20px;margin:0;">
                        {recommendation_items}
                    </ul>
                </td>
            </tr>
        </table>
        """

        # Embed the attached image in the HTML using CID
        cid = "top_coins_plot"  # This should match the Content-ID of the attached image
        plot_html = f"""
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
            <tr>
                <td style="padding:20px;text-align:center;">
                    <h3 style="font-size:20px;color:#2a9d8f;margin-bottom:10px;">Top Coins Cumulative Scores Over Time</h3>
                    <img src="cid:{cid}" alt="Top Coins Plot" style="width:100%;max-width:600px;height:auto;"/>
                </td>
            </tr>
        </table>
        """

    # Full HTML structure
    html_content = f"""
    <html>
    <body style="margin:0;padding:0;background-color:#f9f9f9;font-family:Arial,sans-serif;color:#333;">
        <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:#f9f9f9;">
            <tr>
                <td align="center">
                    <table width="600" cellpadding="0" cellspacing="0" border="0" style="background-color:#fff;">
                        <tr>
                            <td style="padding:20px;">
                                <h2 style="text-align:center;color:#264653;font-size:24px;margin:0;">Coin Analysis Report</h2>
                            </td>
                        </tr>
                        <tr>
                            <td>
                                {digest_html}
                            </td>
                        </tr>
                        <tr>
                            <td>
                                {recommendations_html}
                            </td>
                        </tr>
                        {plot_html}
                        <tr>
                            <td style="padding:20px;">
                                <p style="text-align:center;color:#777;font-size:12px;margin:0;">Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                            </td>
                        </tr>
                    </table>
                </td>
            </tr>
        </table>
    </body>
    </html>
    """

    return html_content

# Number of rows (coins) per batch
ROWS_PER_BATCH = 50

def gpt4o_summarize_batch(batch_df):
    try:
        df_json = batch_df.to_dict(orient='records')
        dataset_json = json.dumps(df_json, indent=2).replace('%', '%%')
    except Exception as e:
        logging.error(f"Failed to serialize batch: {e}")
        return {"recommendations": []}

    prompt = f"""
    You are provided with structured analysis data for multiple cryptocurrency coins.

    Your task is to **summarize the data for each coin** using the available metrics. You are NOT deciding whether to recommend the coin. Instead, describe the coin's performance in intuitive, human-readable language. You should also summarise any news on the coin.

    ---

    ### Interpret the following metrics in your explanation:

    - **liquidity_risk**: Indicates how easily the coin can be traded. Use "low", "medium", or "high".
    - **price_change_score**:
        - 0 = No significant price changes
        - 1 = A notable increase in price in one time window (e.g., short-term)
        - 2 = Strong increases in two time windows
        - 3 = Consistent strong price momentum across short, medium, and long term
    - **volume_change_score**:
        - 0 = No notable increase in trading activity
        - 1 = A moderate spike in trading volume in one time window
        - 2 = Sustained and strong trading activity over multiple periods
    - **cumulative_score**: Combined measure of price, volume, sentiment, trends, and other signals
        - 0–2 = Weak signals overall
        - 3–5 = Moderate momentum
        - 6+ = Strong breakout potential or multi-signal alignment
    - **trend_conflict**: 
        - "Yes" = The coin shows consistent monthly growth but lacks short-term support — this may indicate a potential early-stage breakout or lagging price action.

    Your summary must turn these scores into natural language and explain what they mean for each coin.

    ---

    ### Output Format (Structured JSON Only):

    ```json
    {{
    "recommendations": [
        {{
        "coin": "Coin Name",
        "liquidity_risk": "Low/Medium/High",
        "cumulative_score": "Score Value",
        "recommendation": "Yes",
        "reason": "An intuitive summary using natural language to explain price, volume, sentiment, and liquidity scores. Also summarize any news on the coin."
        }}
    ]
    }}
    ```

    ---

    ### Instructions:

    - Always use \"recommendation\": \"Yes\".
    - If trend_conflict is “Yes”, describe it as a possible early-stage breakout, not a red flag.
    - Do NOT quote raw scores alone — always explain what they mean.
    - Use a fluent, confident tone. Avoid jargon or numbers without meaning.
    - Do NOT summarize the dataset as a whole.
    - Summarise any news provided for the coin.
    - Return only valid structured JSON.

    Here is the dataset:
    {dataset_json}
    """

    def api_call():
        return openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0
        )

    try:
        from api_clients import api_call_with_retries  # Ensure this is imported where needed
        response = api_call_with_retries(api_call)
        content = response['choices'][0]['message']['content']

        match = re.search(r'```json(.*?)```', content, re.DOTALL)
        if match:
            json_data = json.loads(match.group(1).strip())
            return json_data
        else:
            logging.warning("No JSON block found in GPT response.")
            return {"recommendations": []}

    except Exception as e:
        logging.error(f"GPT-4o batch failed: {e}")
        return {"recommendations": []}


def gpt4o_summarize_each_coin(df, rows_per_batch=ROWS_PER_BATCH, num_processes=None):
    """
    Process a DataFrame of coin metrics in parallel, batching by number of rows.

    Args:
        df (pd.DataFrame): Coin metrics.
        rows_per_batch (int): Number of coins per GPT call.
        num_processes (int or None): Number of processes to use (default: auto).

    Returns:
        dict: Combined GPT-generated recommendations for all coins.
    """
    batches = [df.iloc[i:i + rows_per_batch] for i in range(0, len(df), rows_per_batch)]

    if num_processes is None:
        num_processes = min(cpu_count(), len(batches))

    with Pool(num_processes) as pool:
        results = pool.map(gpt4o_summarize_batch, batches)

    all_recommendations = []
    for result in results:
        if result and 'recommendations' in result:
            all_recommendations.extend(result['recommendations'])

    return {"recommendations": all_recommendations}

def gpt4o_analyze_and_recommend(df):
    """
    Uses GPT-4o to analyze the final results DataFrame and provide structured recommendations for coin purchases.

    Parameters:
        df (pd.DataFrame): The final DataFrame containing coin analysis results.

    Returns:
        dict: A structured summary of GPT-4o's recommendations for coin purchases, including reasons.
    """
    # Convert DataFrame to JSON for input
    try:
        df_json = df.to_dict(orient='records')
        dataset_json = json.dumps(df_json, indent=2).replace('%', '%%')
    except Exception as e:
        logging.error(f"Failed to serialize df_json: {e}")
        dataset_json = "{}"

    prompt = f"""
You are provided with structured analysis data for multiple cryptocurrency coins. Your task is to **evaluate each coin individually** and decide if it should be considered for **purchase based on the potential for a breakout or surge in value**.

---

Please follow this **step-by-step reasoning process** for each coin:

### Step-by-Step Evaluation:

1. **Check data completeness and uniqueness**:
   - If the coin record lacks required metrics or is a duplicate (based on name or ID), skip it.

2. **Assess Liquidity Risk**:
   - Use the `liquidity_risk` metric to judge how easily the coin can be bought or sold.
   - High liquidity risk is a red flag unless strongly offset by other metrics.

3. **Evaluate Sentiment and Market Momentum**:
   - Analyze `sentiment_score`. A high score indicates bullish community outlook.
   - Cross-check with `volume` and `price change scores` to confirm rising interest.

4. **Analyze Cumulative Score**:
   - Use the `cumulative_score` as a key signal of overall strength.
   - Consider coins with high cumulative scores and strong support from sentiment/volume as breakout candidates.

5. **Apply Decision Criteria**:
   - Recommend "Yes" only if there is **strong and clear evidence** supporting a breakout potential.
   - Recommend "No" only if there is **clear and confident evidence against** a breakout.

6. **Generate Explanation**:
   - Clearly justify the recommendation using data.
   - Reference at least two specific metrics (e.g., “High cumulative score (89.5), strong sentiment (+0.6), and low liquidity risk suggest bullish potential.”)

---

### Output Format (Structured JSON Only):

```json
{{
  "recommendations": [
    {{
      "coin": "Coin Name",
      "liquidity_risk": "Low/Medium/High",
      "cumulative_score": "Score Value",
      "recommendation": "Yes/No",
      "reason": "A fluent, specific, data-driven explanation referencing key metrics."
    }}
  ]
}}
```

### Additional Rules:

- Do **not** summarize the full dataset.
- Do **not** include coins with unclear or incomplete signals.
- Do **not** include multiple entries for the same coin — deduplicate and select the most recent/relevant one.
- Use a **confident, analytical tone**, not speculative or vague language.
- Your output should **only** include the final structured JSON.

Now, here is the dataset:
{dataset_json}
"""

    def api_call():
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            n=1,
            stop=None,
            temperature=0.0
        )
        return response

    try:
        response = api_call_with_retries(api_call)
        gpt_message_content = response['choices'][0]['message']['content']

        logging.debug("Raw GPT response:\n%s", gpt_message_content)

        json_match = re.search(r'```json(.*?)```', gpt_message_content, re.DOTALL)
        if json_match:
            json_content = json_match.group(1).strip()
            logging.debug("Extracted JSON content:\n%s", json_content)

            parsed_data = json.loads(json_content)
            logging.debug(f"Parsed JSON data: {parsed_data}")
            return parsed_data

        logging.debug("No JSON content found in the GPT response.")
        return {"recommendations": []}

    except Exception as e:
        logging.error(f"Failed to complete GPT-4o analysis: {e}")
        return {"recommendations": []}
    
def send_failure_email():
    """
    Sends an email with the current results when the script encounters an error.
    If no flag file for the current date exists, it deletes all previous flag files,
    sends the email, and creates a flag file for today.
    """

    # Get today's date in the format YYYY-MM-DD
    today = datetime.now().strftime("%Y-%m-%d")

    # Define the flag file path for today's date
    flag_file = f"email_sent_{today}.flag"

    # Check if the flag file for today already exists
    if os.path.exists(flag_file):
        logging.debug(f"Email already sent today ({today}). Skipping email.")
        return  # Exit the function if email has already been sent today

    # Delete all previous flag files if today's flag file does not exist
    flag_files = glob.glob("email_sent_*.flag")
    for file in flag_files:
        try:
            os.remove(file)
            logging.debug(f"Deleted old flag file: {file}")
        except Exception as e:
            logging.debug(f"Failed to delete flag file {file}: {e}")

    # Proceed to send the email if no flag file exists for today
    if os.path.exists(LOG_DIR+'/coin_analysis_report.xlsx'):
        with open(LOG_DIR+'/coin_analysis_report.xlsx', 'r') as file:
            file_contents = file.read()
    else:
        file_contents = "No data available, as the results file was not created."

    # HTML content with inline CSS for the failure email
    html_content = f"""
    <html>
    <head>
        <style>
            body {{
                font-family: Arial, sans-serif;
                color: #333;
            }}
            h2 {{
                color: #c0392b;
            }}
            p {{
                font-size: 14px;
                color: #555;
            }}
            .content {{
                background-color: #f9f9f9;
                padding: 20px;
                border: 1px solid #ddd;
                border-radius: 5px;
            }}
            .content pre {{
                background-color: #f4f4f4;
                border: 1px solid #ccc;
                padding: 10px;
                border-radius: 3px;
            }}
        </style>
    </head>
    <body>
        <h2>Failure in Weekly Coin Analysis Script</h2>
        <p>The script encountered an error. Below are the current results:</p>
        <div class="content">
            <pre>{file_contents}</pre>
        </div>
    </body>
    </html>
    """

    msg = MIMEMultipart('alternative')
    msg['Subject'] = "Failure in Weekly Coin Analysis Script"
    msg['From'] = EMAIL_FROM
    #msg['To'] = EMAIL_TO
    msg['Bcc'] = EMAIL_TO   # Add BCC field (replace EMAIL_BCC with your BCC email address)

    part = MIMEText(html_content, 'html')
    msg.attach(part)

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            recipients = EMAIL_TO.split(",")  # Add BCC recipients to the send list
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        logging.debug("Failure email sent successfully.")

        # Create today's flag file to avoid sending multiple emails
        with open(flag_file, 'w') as f:
            f.write("Email sent")

    except Exception as e:
        logging.debug(f"Failed to send email: {e}")


def print_command_line_report(report_entries):
    """
    Prints a command-line report of the daily coin analysis.

    Parameters
    ----------
    report_entries : list
        A list of dictionaries, each containing the analysis results for a single coin.

    Returns
    -------
    None
    """
    df = pd.DataFrame(report_entries)
    logging.debug("\nCoin Analysis Report")
    logging.debug(tabulate(df, headers="keys", tablefmt="grid"))
    logging.debug(f"\nReport generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

def gpt4o_summarize_digest_and_extract_tickers(digest_text):
    """
    Uses GPT-4 to summarize the Sundown Digest and extract key points related to potential surges in coin value.

    Args:
        digest_text (str): The concatenated text from all digest entries.

    Returns:
        dict: A dictionary containing a summary focused on surge-causing news and a list of extracted tickers.
    """
    prompt = f"""
    Analyze the following digest entries and provide the following:
    1) A concise summary in bullet points (no more than 250 words) of key news items likely to cause surges in the value of the mentioned coins. 
    2) List the relevant cryptocurrency tickers beside each news item. Ensure there is no duplication.

    Text:
    {digest_text}

    Respond **only** in JSON format with 'surge_summary' and 'tickers' as keys. Ensure the tickers are in alphabetical order and there are no duplicate tickers.
    """

    try:
        response = openai.ChatCompletion.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            n=1,
            stop=None,
            temperature=0.0
        )
        
        # Extract the content of the response
        response_content = response.choices[0].message['content'].strip()
        # Use a regular expression to extract JSON from the response content
        json_match = re.search(r'\{.*\}', response_content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0)
            try:
                analysis = json.loads(json_str)
                return analysis
            except json.JSONDecodeError:
                logging.debug(f"Failed to decode JSON: {json_str}")
                return {"surge_summary": "", "tickers": []}
        else:
            logging.debug(f"No JSON found in the response: {response_content}")
            return {"surge_summary": "", "tickers": []}
        
    except openai.error.RateLimitError as e:
        logging.debug(f"Rate limit reached: {e}. Waiting for 60 seconds before retrying...")
        time.sleep(60)  # Wait before retrying
        return gpt4o_summarize_digest_and_extract_tickers(digest_text)  # Retry the request

    except Exception as e:
        logging.debug(f"An error occurred while summarizing the digest and extracting tickers: {e}")
        return {"surge_summary": "", "tickers": []}


def summarize_sundown_digest(digest):
    """
    Summarizes the Sundown Digest content from the last three days, including sentiment detection,
    coin ticker extraction, and a summary focused on news likely to cause surges in coin value.

    Args:
        digest (list): List of Sundown Digest entries.

    Returns:
        dict: A dictionary containing key points of news that may cause surges and relevant tickers.
    """
    # Get the current date and calculate the date three days ago
    current_date = datetime.now()
    three_days_ago = current_date - timedelta(days=3)

    digest_texts = []
    
    for entry in digest:
        # Parse the entry's date
        entry_date = datetime.strptime(entry['date'], '%Y-%m-%dT%H:%M:%S.%fZ')

        # Filter out entries older than three days
        if entry_date < three_days_ago:
            continue

        digest_texts.append(entry['text'])

    # Concatenate all digest texts into a single string
    combined_digest_text = " ".join(digest_texts)

    # Use GPT-4 to analyze and summarize the combined digest text
    summary_and_tickers = gpt4o_summarize_digest_and_extract_tickers(combined_digest_text)
    return summary_and_tickers

def send_email_with_report(html_content, attachment_path, plot_image_path=LOG_DIR+'top_coins_plot.png', recommendations=None):
    """
    Sends an email with an HTML report and an attached image.

    The email uses a 'related' MIME type to allow both HTML and images to be attached.
    The HTML content is passed as a string and the image is attached as an inline
    attachment with a Content-ID header that matches the CID in the HTML content.

    Args:
        html_content (str): The HTML content of the email.
        attachment_path (str): The path to the Excel file attachment.
        plot_image_path (str): The path to the image file to attach.
        recommendations (list): List of recommended coins, if any.

    Returns:
        None
    """
    try:
        logging.debug(f"Preparing email with attachment: {attachment_path} and plot: {plot_image_path}")
        
        if not EMAIL_FROM or not EMAIL_TO:
            logging.error("EMAIL_FROM is not set. Cannot send email.")
            return
        
        # Create a 'related' multipart message for HTML + images
        msg = MIMEMultipart('related')  # 'related' allows attaching both HTML and images
        msg['Subject'] = "AI Generated Coin Analysis Report"
        msg['From'] = EMAIL_FROM
        #msg['To'] = EMAIL_TO
        msg['Bcc'] = EMAIL_TO  # Add BCC field (replace EMAIL_BCC with your BCC email address)

        # Attach HTML content
        part = MIMEText(html_content, 'html')
        msg.attach(part)

        # Conditionally attach the plot only if there are recommendations
        if recommendations and len(recommendations) > 0:
            logging.debug(f"Attaching plot image: {plot_image_path}")
            try:
                with open(plot_image_path, 'rb') as img_file:
                    mime_image = MIMEImage(img_file.read(), _subtype='png')
                    mime_image.add_header('Content-ID', '<top_coins_plot>')  # Content-ID should match CID in HTML
                    mime_image.add_header('Content-Disposition', 'inline', filename=LOG_DIR+"/top_coins_plot.png")
                    msg.attach(mime_image)
            except Exception as e:
                logging.error(f"Error attaching plot image: {e}")
                logging.debug(traceback.format_exc())  # Log the full stack trace

        # Attach the Excel file
        if os.path.exists(attachment_path):
            logging.debug(f"Attaching Excel file: {attachment_path}")
            try:
                with open(attachment_path, 'rb') as file:
                    part = MIMEApplication(file.read(), _subtype="xlsx")
                    part.add_header('Content-Disposition', 'attachment', filename=os.path.basename(attachment_path))
                    msg.attach(part)
            except Exception as e:
                logging.error(f"Error attaching Excel file: {e}")
                logging.debug(traceback.format_exc())  # Log the full stack trace


        try:
            logging.debug(f"Connecting to SMTP server: {SMTP_SERVER}")
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
                recipients = EMAIL_TO.split(",")
                server.sendmail(EMAIL_FROM, recipients, msg.as_string())
            logging.debug("Email sent successfully.")  # <-- Only log here, inside try
        except Exception as e:
            logging.error(f"Error sending email: {e}")
            logging.debug(traceback.format_exc())  # Log the full stack trace


    except Exception as e:
        logging.error(f"An error occurred in send_email_with_report: {e}")
        logging.debug(traceback.format_exc())  # Log the full stack trace for debugging



def save_report_to_excel(report_entries, filename=LOG_DIR+'/coin_analysis_report.xlsx'):
    """
    Saves the report entries to an Excel file with enhanced formatting and styling.

    Args:
        report_entries (list): A list of dictionaries containing the report data.
        filename (str): The name of the Excel file to save the report to.
    """
    # Convert the report entries to a pandas DataFrame
    df = pd.DataFrame(report_entries)
    
    # Save DataFrame to an Excel file without formatting
    df.to_excel(filename, index=False)
    
    # Open the Excel file with openpyxl for formatting
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Define styles for headers and cells
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    cell_font = Font(name="Arial", size=10)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)  # Turn off wrap_text for content cells
    
    # Define border style
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Apply header styles (background color, font, alignment)
    for col in sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
        max_length = 0
        column = col[0].column_letter  # Get the column letter for header
        for cell in col:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)  # Turn wrapping off for headers
            cell.border = thin_border
            # Adjust column width based on header content
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))

        adjusted_width = (max_length + 2) * 1.2  # Add some padding for headers
        sheet.column_dimensions[column].width = adjusted_width

    # Apply cell styles (font, alignment, borders) and auto-adjust column width based on content
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length = 0
        column = col[0].column_letter  # Get the column letter for data cells

        for cell in col:
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Adjust column width based on the content
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except Exception as e:
                print(f"Error processing cell {cell.coordinate}: {e}")

        # Set the column width to fit the content with padding
        adjusted_width = (max_length + 2) * 1.2  # Add padding for cells
        sheet.column_dimensions[column].width = adjusted_width

    # Freeze the top row (headers) for better readability
    sheet.freeze_panes = "A2"

    # Save the workbook with the formatting applied
    try:
        workbook.save(filename)
        print(f"Report saved to {filename} with enhanced formatting.")
    except Exception as e:
        print(f"Error saving the report: {e}")
    finally:
        workbook.close()

    return filename
